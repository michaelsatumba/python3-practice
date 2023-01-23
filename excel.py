import openpyxl
import os
#print(os.getcwd())

'''
workbook = openpyxl.load_workbook('example.xlsx')
print(workbook)
sheet = workbook.get_sheet_by_name('Sheet1')
print(sheet)
print(workbook.get_sheet_names())
cell = sheet['A1']
print(cell.value)
'''

wb = openpyxl.Workbook()
print(wb)
print(wb.get_sheet_names())
sheet = wb.get_sheet_by_name('Sheet')
sheet['A1'] = 42
sheet.title = 'English'

sheet2 = wb.create_sheet()
sheet2.title = 'Spanish'

sheet2['A1'] = 'hola'

wb.create_sheet(index=0, title='My first Sheet')
fruits = ['apple', 'orange']

for i, fruit in enumerate(fruits):
    sheet2.cell(row=i+1, column=2).value = fruit


wb.save('exampleOne.xlsx')
